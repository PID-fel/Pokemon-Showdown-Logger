from datetime import datetime
from xlwt import Workbook 
from openpyxl import Workbook
from openpyxl import load_workbook
import os
from operator import itemgetter
import shutil
import gspread
import json
import pathlib

sheetName = "showdown.xlsx"
accounts = []

downloadsPathXLSX = "./To_Log_XLSX_Replays/"
downloadsPathGoogleSheets = "./To_Log_Google_Sheet_Replays/"


loggedGamesPath = "./Logged_Replays/"

data = json.loads(pathlib.Path('credentials.json').read_text())

scopes = ["https://www.googleapis.com/auth/spreadsheets"]
sheetId = "1nPH6Csv5pRwMmE8mChgHhecHzwsqAXMYecNIuMByKI4"

gc = gspread.service_account_from_dict(data)

def getAccounts(accountsFileName):
    with open(accountsFileName) as f:
        allLines = f.readlines()
        if len(allLines) == 0:
            raise ValueError("No User Accounts Specified in Accounts")
        else:
            return(allLines)
    
def intToColumnLetter(int):
    #Only does up to two digits which isn't ideal but I doubt you would ever have more than 26*27 columns

    alphabet = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O', 'P', 'Q', 'R', 'S', 'T', 'U', 'V', 'W', 'X', 'Y', 'Z']
    begining = ''
    end = alphabet[int % 26]

    if int >= 26:
        begining = alphabet[(int // 26) - 1]

    return begining + end

def logDictionaryToInputList(logDictionary):
    outList = []
    dictionaryKeys = list(logDictionary.keys())
    for x in dictionaryKeys:
        if not (x == "p1PokemonList" or x == "p2PokemonList"):
            outList.append(logDictionary[x])

    return outList

def addHeaderToSheet(fileName):
    columnHeaders = ["Index", "Filename", "Format", "Player1", "Player2",
        "P1Poke0", "P1Poke1" , "P1Poke2" , "P1Poke3", "P1Poke4", "P1Poke5",
        "P2Poke0", "P2Poke1" , "P2Poke2" , "P2Poke3", "P2Poke4", "P2Poke5",
        "P1Elo_Start", "P1Elo_End", "P2Elo_Start", "P2Elo_End",
        "Unix_Timestamp_Start", "Date_Start", "Time_Start", "Time_End", "Total_Turns"]
    workbook = load_workbook(filename=fileName)
    workbook.sheetnames
    sheet = workbook.active
    columnIndex = 0

    for columnHeader in columnHeaders:
        sheet[intToColumnLetter(columnIndex) + "1"] =  columnHeader
        columnIndex = columnIndex + 1
    workbook.save(filename=fileName)

def gameLogTodictionary(fileName, accountList):

    outList = None

    with open(fileName, encoding="utf-8", errors='ignore') as f:
        outList = f.read().split('\n')

    gameLogDictionary = {
        "fileName" : fileName.split("/")[-1],
        "format" : (fileName.split("/")[-1].split("-")[0]),
        "p1": None,
        "p2": None,
        "p1PokemonList": [],
        "p1Poke0" : None, "p1Poke1" : None, "p1Poke2" : None, "p1Poke3" : None, "p1Poke4" : None, "p1Poke5" : None,
        "p2Poke0" : None, "p2Poke1" : None, "p2Poke2" : None, "p2Poke3" : None, "p2Poke4" : None, "p2Poke5" : None,
        "p2PokemonList": [],
        "p1RatingStart": None,
        "p1RatingEnd": None,
        "p2RatingStart": None,
        "p2RatingEnd": None,
        "dateTimeStart" : None,
    }

    p1revealedPokemon = []
    p2revealedPokemon = []
    allTimes = []
    turnTimes = []
    unrevealedForms1 = []
    unrevealedForms2 = []

    for x in outList:
        if x[0:6] == "|poke|":
            pokemonName = x[9:].split(',')[0].split('|')[0]
            if x[6:8] == "p1":
                gameLogDictionary["p1PokemonList"].append(pokemonName)
                if pokemonName.split("-")[-1] == "*":
                    unrevealedForms1.append(pokemonName)

            if x[6:8] == "p2":
                gameLogDictionary["p2PokemonList"].append(pokemonName)
                if pokemonName.split("-")[-1] == "*":
                    unrevealedForms2.append(pokemonName)

        if x[0:4] == "|t:|":
            allTimes.append(x[4:])
        
        if x[0:6] == "|turn|":
            turnTimes.append(allTimes[-1])

        if x[0:5] == "|raw|" and not x[0:9] == "|raw|<div" and not x[0:10] == "|raw|<font": 
            ratingSplit = x.split("rating:")
            if len(ratingSplit) >1:
                
                endRating = ratingSplit[1].split("<")[1].split(">")[1]
                if gameLogDictionary["p1RatingEnd"] == None:
                    gameLogDictionary["p1RatingEnd"] = endRating
                else:
                    gameLogDictionary["p2RatingEnd"] = endRating

        if x[0:8] == "|player|":
            if len(x)>16:
                rating = x[-4:]
                playerName = x[11:].split("|")[0]

                if x[0:11] == "|player|p1|" and gameLogDictionary["p1"] == None:
                    gameLogDictionary["p1"] = playerName
                    gameLogDictionary["p1RatingStart"] = rating
                elif x[0:11] == "|player|p2|" and gameLogDictionary["p2"] == None:
                    gameLogDictionary["p2"] = playerName
                    gameLogDictionary["p2RatingStart"] = rating

        if x[0:9] == "|switch|p":
            player = x[9:10]
            revealedpokemon = x.split(": ")[1].split("|")[1].split(",")[0]
            if player == "1" and (revealedpokemon not in p1revealedPokemon):
                p1revealedPokemon.append(revealedpokemon)

            if player == "2" and (revealedpokemon not in p2revealedPokemon):
                p2revealedPokemon.append(revealedpokemon)

    gameLogDictionary["date"] = datetime.fromtimestamp(int(allTimes[0])).strftime("%Y-%m-%d") 
    gameLogDictionary["timeStart"] = datetime.fromtimestamp(int(allTimes[0])).strftime("%H:%M:%S") 
    gameLogDictionary["timeFinish"] = datetime.fromtimestamp(int(allTimes[-1])).strftime("%H:%M:%S") 
    gameLogDictionary["dateTimeStart"] = int(allTimes[-1])
    gameLogDictionary["turnCount"] = len(turnTimes)

    if len(gameLogDictionary["p1PokemonList"]) == 0:
        gameLogDictionary["p1PokemonList"] = p1revealedPokemon

    if len(gameLogDictionary["p2PokemonList"]) == 0:
        gameLogDictionary["p2PokemonList"] = p2revealedPokemon

    if any(x.lower().rstrip() == gameLogDictionary["p1"].lower().rstrip() for x in accountList):
        pass
    elif any(x.lower().rstrip() == gameLogDictionary["p2"].lower().rstrip() for x in accountList):

        p1Save = gameLogDictionary["p1"]
        p1RatingStartSave = gameLogDictionary["p1RatingStart"]
        p1RatingEndSave = gameLogDictionary["p1RatingEnd"]
        p1PokemonListSave = gameLogDictionary["p1PokemonList"]
        p1revealedPokemonSave = p1revealedPokemon

        gameLogDictionary["p1"] = gameLogDictionary["p2"]
        gameLogDictionary["p1RatingStart"] = gameLogDictionary["p2RatingStart"]
        gameLogDictionary["p1RatingEnd"] = gameLogDictionary["p2RatingEnd"]
        gameLogDictionary["p1PokemonList"] = gameLogDictionary["p2PokemonList"]
        p1revealedPokemon = p2revealedPokemon

        gameLogDictionary["p2"] = p1Save
        gameLogDictionary["p2RatingStart"] = p1RatingStartSave
        gameLogDictionary["p2RatingEnd"] = p1RatingEndSave
        gameLogDictionary["p2PokemonList"] = p1PokemonListSave
        p2revealedPokemon = p1revealedPokemonSave
        
    else:
        raise ValueError("no username matches:", gameLogDictionary["p1"].lower().rstrip(), "or", gameLogDictionary["p2"].lower().rstrip())

    teamKeys = ["Poke" + str(x) for x in range(6)]


    #changes unrevealed forms into the revealed forms (when possible)
    revealedPokemon = [p1revealedPokemon, p2revealedPokemon]
    for playerNumber in (range(1, 3)):
        for unrevealedFormIndex in range(len(gameLogDictionary["p"+str(playerNumber)+"PokemonList"])):
            unrevealedForm = gameLogDictionary["p"+str(playerNumber)+"PokemonList"][unrevealedFormIndex]
            if unrevealedForm.split("-")[-1] == "*":

                for revealedFormIndex in range (len(revealedPokemon[playerNumber-1])):
                    revealedForm = revealedPokemon[playerNumber-1][revealedFormIndex]

                    if (revealedForm.split("-")[0] == unrevealedForm.split("-")[0]):
                        gameLogDictionary["p"+str(playerNumber)+"PokemonList"][unrevealedFormIndex] = revealedForm
                        revealedPokemon[playerNumber-1][revealedFormIndex] = ""
                        break

    #adds list of pokemon into individual key value pairs ie p1pokemon list:[pikachu, charizard] -> p1Poke0: pikachu, p1poke1, charizard
    #this is for easier access to the names in methods that uses the dictionary
    for x in range (2):
        for y in range (len(teamKeys)):
            currentPlayerPokeList = gameLogDictionary["p"+str(x+1)+"PokemonList"]

            if len(currentPlayerPokeList) <= y:
                gameLogDictionary["p"+str(x+1)+str(teamKeys[y])] = "N/A"
            else:
                gameLogDictionary["p"+str(x+1)+str(teamKeys[y])] = currentPlayerPokeList[y]


    return(gameLogDictionary)

def getListsOfAllGames(gameDownloadsPath):
    dir_list = os.listdir(gameDownloadsPath)

    allGameLogDictionaries = []

    dateTimeStartIndex = None
    dateTimeStartIndex
    for gameName in dir_list:
        allGameLogDictionaries.append(gameLogTodictionary(gameDownloadsPath+gameName, accounts))

    allGameLogDictionariesSorted = sorted(allGameLogDictionaries, key=itemgetter('dateTimeStart'))

    return allGameLogDictionariesSorted

def addListOfGamesToXLSXSheet(allGameLogDictionariesSorted):
    rowToCheck = 2
    unfilledRow = None

    workbook = load_workbook(filename=sheetName)
    sheet = workbook.active

    while unfilledRow == None:

        if (sheet["A"+str(rowToCheck)].value == None):
            unfilledRow = rowToCheck
        else:
            rowToCheck += 1
    
    convertedLists = []
    currentListIndex = 1
    for gameDictionary in allGameLogDictionariesSorted:

        gameDictionary.pop("p1PokemonList", any)
        gameDictionary.pop("p2PokemonList", any)
        convertedList = list(gameDictionary.values())
        convertedList.insert(0, sheet["A"+str(rowToCheck-1)].value + currentListIndex)
        convertedLists.append(convertedList)
        currentListIndex += 1

    for currentRow in range(len(convertedLists)):
        for currentCollumn in range(len(convertedLists[0])):
            cellColumn = intToColumnLetter(currentCollumn)
            cellRow = str(unfilledRow + currentRow)
            sheet[cellColumn+cellRow] = convertedLists[currentRow][currentCollumn]
        gameName = allGameLogDictionariesSorted[currentRow]["fileName"]

        print(gameName, "  added to XLSX")
        shutil.copyfile(downloadsPathXLSX + gameName, downloadsPathGoogleSheets + gameName)
        os.remove(downloadsPathXLSX + gameName)

    workbook.save(filename=sheetName)

def addListOfGamesToGoogleSheet(allGameLogDictionariesSorted):
    sheet = gc.open_by_key(sheetId).sheet1

    AAColumn = sheet.col_values(1)
    firstCellYValue = len(AAColumn) + 1

    convertedLists = []
    currentListIndex = 1
    for gameDictionary in allGameLogDictionariesSorted:
        gameDictionary.pop("p1PokemonList", any)
        gameDictionary.pop("p2PokemonList", any)
        
        convertedList = list(gameDictionary.values())

        convertedList.insert(0, int(AAColumn[-1]) + currentListIndex)
        convertedLists.append(convertedList)
        currentListIndex += 1

    cellBeginning = "A"+str(firstCellYValue)
    cellEnd =str(intToColumnLetter(len(convertedLists[0])-1))+str(firstCellYValue+len(convertedLists)-1)
    sheet.update(cellBeginning+":"+cellEnd, convertedLists)

    for gameDictionary in allGameLogDictionariesSorted:
        gameName = gameDictionary["fileName"]
        print(gameName, "  added to Google Sheet")

        shutil.copyfile(downloadsPathGoogleSheets + gameName, loggedGamesPath + gameName)
        os.remove(downloadsPathGoogleSheets + gameName)

accounts = getAccounts('./accounts.txt')

addListOfGamesToXLSXSheet(getListsOfAllGames(downloadsPathXLSX))

addListOfGamesToGoogleSheet(getListsOfAllGames(downloadsPathGoogleSheets))